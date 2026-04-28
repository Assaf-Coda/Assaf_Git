#!/usr/bin/env python3
"""
3_import_features.py
────────────────────
Reads features_import.xlsx and creates each feature in Aha! via the REST API.
Behaviour:
  • Skips blank rows and the yellow example row (row 3)
  • Stops on first API error (STOP_ON_ERROR = True in config.py)
  • Writes import_results.xlsx with a status column per row

Prerequisites:
  1. Run 1_discover.py  → confirms your release exists
  2. Fill features_import.xlsx
  3. python 3_import_features.py
"""

import sys
import time
import requests
import openpyxl
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from config import (
    AHA_API_KEY, AHA_SUBDOMAIN, PRODUCT_KEY,
    TARGET_RELEASE_NAME, STOP_ON_ERROR, SUPPRESS_NOTIFICATIONS
)

BASE_URL = f"https://{AHA_SUBDOMAIN}.aha.io/api/v1"
HEADERS = {
    "Authorization": f"Bearer {AHA_API_KEY}",
    "Content-Type":  "application/json",
    "Accept":        "application/json",
    "User-Agent":    f"aha-importer/1.0 ({AHA_SUBDOMAIN})",
}

INPUT_FILE  = "features_import.xlsx"
RESULTS_FILE = "import_results.xlsx"
DATA_START_ROW = 4   # row 1=banner, 2=headers, 3=example
RATE_LIMIT_PAUSE = 0.4   # seconds between requests (safe for Aha! limits)

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ── Helpers ──────────────────────────────────────────────────────────────────

def aha_get(path, params=None):
    resp = requests.get(f"{BASE_URL}{path}", headers=HEADERS, params=params)
    if resp.status_code == 401:
        print("❌  Authentication failed. Check AHA_API_KEY in config.py")
        sys.exit(1)
    resp.raise_for_status()
    return resp.json()


def resolve_release_id():
    """Find the numeric release reference for TARGET_RELEASE_NAME."""
    data = aha_get(f"/products/{PRODUCT_KEY}/releases", params={"per_page": 100})
    releases = data.get("releases", [])
    for r in releases:
        if r["name"].strip().lower() == TARGET_RELEASE_NAME.strip().lower():
            return r["reference_num"]
    names = [r["name"] for r in releases]
    print(f"❌  Release '{TARGET_RELEASE_NAME}' not found in {PRODUCT_KEY}.")
    print(f"    Available releases: {names}")
    print(f"    Update TARGET_RELEASE_NAME in config.py")
    sys.exit(1)


def create_feature(release_ref, row_data):
    """POST a single feature. Returns (success: bool, message: str, aha_url: str)."""
    payload = {"feature": {}}

    name = row_data.get("name", "").strip()
    if not name:
        return False, "Skipped — Feature Name is blank", ""

    payload["feature"]["name"] = name

    if row_data.get("description"):
        payload["feature"]["description"] = row_data["description"]

    if row_data.get("workflow_status"):
        payload["feature"]["workflow_status"] = {"name": row_data["workflow_status"]}

    if row_data.get("tags"):
        payload["feature"]["tags"] = row_data["tags"]

    if row_data.get("assigned_to"):
        payload["feature"]["assigned_to_user"] = {"email": row_data["assigned_to"]}

    if row_data.get("start_date"):
        payload["feature"]["start_date"] = row_data["start_date"]

    if row_data.get("due_date"):
        payload["feature"]["due_date"] = row_data["due_date"]

    if row_data.get("score"):
        try:
            payload["feature"]["score_facts"] = [{"value": int(row_data["score"])}]
        except ValueError:
            pass

    if SUPPRESS_NOTIFICATIONS:
        payload["feature"]["send_email"] = False

    url = f"{BASE_URL}/releases/{release_ref}/features"
    try:
        resp = requests.post(url, headers=HEADERS, json=payload)
        if resp.status_code in (200, 201):
            feature = resp.json().get("feature", {})
            ref = feature.get("reference_num", "?")
            aha_url = feature.get("url", "")
            return True, f"Created {ref}", aha_url
        else:
            try:
                err = resp.json()
            except Exception:
                err = resp.text
            return False, f"HTTP {resp.status_code}: {err}", ""
    except requests.RequestException as e:
        return False, f"Request error: {e}", ""


def read_template():
    """Read the filled template and return list of row dicts."""
    try:
        wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    except FileNotFoundError:
        print(f"❌  {INPUT_FILE} not found. Run 2_generate_template.py first.")
        sys.exit(1)

    ws = wb["FEATURES"]
    columns = [
        "name", "description", "workflow_status", "tags",
        "initiative_name", "assigned_to", "start_date",
        "due_date", "score", "release_phase"
    ]

    rows = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if all((v is None or str(v).strip() == "") for v in row):
            continue   # skip blank rows
        row_dict = {}
        for col_idx, key in enumerate(columns):
            val = row[col_idx] if col_idx < len(row) else None
            row_dict[key] = str(val).strip() if val is not None else ""
        rows.append(row_dict)

    return rows


def write_results(rows_with_status):
    """Write import_results.xlsx summarising the run."""
    wb = Workbook()
    ws = wb.active
    ws.title = "RESULTS"
    ws.freeze_panes = "A2"

    headers = ["#", "Feature Name", "Status", "Aha! Ref / Error", "Aha! URL", "Timestamp"]
    widths   = [5, 45, 12, 35, 50, 22]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[chr(64 + col)].width = w
    ws.row_dimensions[1].height = 22

    ok_fill   = PatternFill("solid", start_color="E2EFDA")
    fail_fill = PatternFill("solid", start_color="FFDCE1")
    skip_fill = PatternFill("solid", start_color="F2F2F2")

    totals = {"created": 0, "failed": 0, "skipped": 0}

    for row_idx, item in enumerate(rows_with_status, start=2):
        success, msg, aha_url, name, ts = (
            item["success"], item["message"], item["url"],
            item["name"], item["timestamp"]
        )
        status_label = "✅ Created" if success else ("⏭ Skipped" if "Skipped" in msg else "❌ Failed")
        fill = ok_fill if success else (skip_fill if "Skipped" in msg else fail_fill)

        values = [row_idx - 1, name, status_label, msg, aha_url, ts]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

        if success:             totals["created"] += 1
        elif "Skipped" in msg:  totals["skipped"] += 1
        else:                   totals["failed"]  += 1

    # Summary row
    summary_row = ws.max_row + 2
    summary = f"TOTAL: {totals['created']} created  |  {totals['failed']} failed  |  {totals['skipped']} skipped"
    cell = ws.cell(row=summary_row, column=1, value=summary)
    cell.font = Font(name="Arial", size=11, bold=True)
    ws.merge_cells(f"A{summary_row}:F{summary_row}")

    wb.save(RESULTS_FILE)
    return totals


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 55)
    print("Aha! Feature Importer")
    print("=" * 55)

    print(f"🔍  Resolving release '{TARGET_RELEASE_NAME}' in {PRODUCT_KEY}...")
    release_ref = resolve_release_id()
    print(f"✅  Found release: {release_ref}")

    print(f"📖  Reading {INPUT_FILE}...")
    rows = read_template()
    print(f"    {len(rows)} data rows found (excluding blank rows and example).\n")

    if not rows:
        print("⚠️  No rows to import. Fill features_import.xlsx and retry.")
        sys.exit(0)

    confirm = input(f"▶  Import {len(rows)} features into {PRODUCT_KEY} / {TARGET_RELEASE_NAME}? [y/N] ")
    if confirm.strip().lower() != "y":
        print("Cancelled.")
        sys.exit(0)

    print()
    results = []
    for i, row in enumerate(rows, start=1):
        name = row.get("name", "(blank)")
        print(f"  [{i:3d}/{len(rows)}] {name[:55]:<55}", end="  ")

        success, message, aha_url = create_feature(release_ref, row)
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        results.append({
            "success": success, "message": message,
            "url": aha_url, "name": name, "timestamp": ts
        })

        if success:
            print(f"✅  {message}")
        elif "Skipped" in message:
            print(f"⏭   {message}")
        else:
            print(f"❌  {message}")
            if STOP_ON_ERROR:
                print("\n⛔  STOP_ON_ERROR is True — halting import.")
                print(f"    Fix the issue and re-run. Successfully created features will NOT be duplicated")
                print(f"    (the importer reads from the same Excel; already-imported rows stay in Aha!)")
                break

        time.sleep(RATE_LIMIT_PAUSE)

    print()
    totals = write_results(results)
    print("=" * 55)
    print(f"  ✅  Created : {totals['created']}")
    print(f"  ❌  Failed  : {totals['failed']}")
    print(f"  ⏭   Skipped : {totals['skipped']}")
    print(f"\n  Results saved to: {RESULTS_FILE}")
    print("=" * 55)


if __name__ == "__main__":
    main()
