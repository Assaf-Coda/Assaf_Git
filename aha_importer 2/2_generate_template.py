#!/usr/bin/env python3
"""
2_generate_template.py
──────────────────────
Reads discovery_summary.json and generates features_import.xlsx.

Sheet 1 — FEATURES   : one row per feature to import
Sheet 2 — INSTRUCTIONS: field-by-field guidance
Sheet 3 — _LOOKUPS   : hidden data for dropdowns (do not edit)
"""

import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Colours ──────────────────────────────────────────────────────────────────
COL_HEADER_REQUIRED  = "1F4E79"   # dark blue   – mandatory field
COL_HEADER_OPTIONAL  = "2E75B6"   # mid blue    – optional field
COL_HEADER_META      = "595959"   # grey        – do not edit column
COL_HEADER_FONT      = "FFFFFF"
COL_ROW_EVEN         = "EBF3FB"
COL_ROW_ODD          = "FFFFFF"
COL_EXAMPLE          = "FFF2CC"   # yellow tint – example row

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Field definitions ────────────────────────────────────────────────────────
# (key, label, required, width, notes)
FIELDS = [
    ("name",             "Feature Name *",        True,  45, "Short, descriptive title"),
    ("description",      "Description",           False, 60, "Plain text. HTML not required."),
    ("workflow_status",  "Workflow Status",        False, 22, "Leave blank for default ('New')"),
    ("tags",             "Tags",                  False, 30, "Comma-separated, e.g. mobile,api"),
    ("initiative_name",  "Initiative / Theme",    False, 30, "Name of initiative to link (optional)"),
    ("assigned_to",      "Assigned To (email)",   False, 30, "Must be a valid Aha! user email"),
    ("start_date",       "Start Date",            False, 15, "YYYY-MM-DD"),
    ("due_date",         "Due Date",              False, 15, "YYYY-MM-DD"),
    ("score",            "Score / Priority",      False, 12, "Numeric — Aha! score value"),
    ("release_phase",    "Release Phase",         False, 20, "Optional release phase name"),
]

EXAMPLE_ROW = [
    "Support multi-tenant SSO login",
    "As an enterprise admin I want to configure SSO so that users can log in via their IdP.",
    "Under consideration",
    "security,enterprise",
    "",
    "",
    "2027-01-15",
    "2027-03-31",
    "70",
    "",
]


def load_summary():
    path = Path("discovery_summary.json")
    if not path.exists():
        print("❌  discovery_summary.json not found.")
        print("    Run 1_discover.py first.")
        sys.exit(1)
    with open(path) as f:
        return json.load(f)


def style_header_cell(cell, required=True):
    bg = COL_HEADER_REQUIRED if required else COL_HEADER_OPTIONAL
    cell.font = Font(bold=True, color=COL_HEADER_FONT, size=11, name="Arial")
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border


def style_data_cell(cell, row_idx):
    bg = COL_ROW_EVEN if row_idx % 2 == 0 else COL_ROW_ODD
    cell.fill = PatternFill("solid", start_color=bg)
    cell.font = Font(name="Arial", size=10)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = border


def build_features_sheet(wb, summary):
    ws = wb.active
    ws.title = "FEATURES"
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 36

    # Row 1 — product info banner
    ws.merge_cells(f"A1:{get_column_letter(len(FIELDS))}1")
    banner = ws["A1"]
    banner.value = (
        f"Aha! Feature Import  ·  Product: {summary['product_key']}  ·  "
        f"Target Release: all features will be created in the release configured in config.py"
    )
    banner.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    banner.fill = PatternFill("solid", start_color="1F4E79")
    banner.alignment = Alignment(horizontal="left", vertical="center")

    # Row 2 — column headers
    for col_idx, (key, label, required, width, _) in enumerate(FIELDS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        style_header_cell(cell, required=required)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Example row (row 3, yellow tint)
    for col_idx, value in enumerate(EXAMPLE_ROW, start=1):
        cell = ws.cell(row=3, column=col_idx, value=value)
        cell.fill = PatternFill("solid", start_color=COL_EXAMPLE)
        cell.font = Font(name="Arial", size=10, italic=True, color="595959")
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border

    # Label the example row
    ws.cell(row=3, column=1).comment = None  # reset
    # Add 47 blank data rows (rows 4–50)
    for row_idx in range(4, 51):
        for col_idx in range(1, len(FIELDS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx, value="")
            style_data_cell(cell, row_idx)

    # ── Workflow status dropdown ─────────────────────────────────────────────
    statuses = [
        "New", "Under consideration", "Planned", "In development",
        "Shipped", "Won't implement"
    ]
    # Write to _LOOKUPS sheet and reference from there
    lookup_ws = wb["_LOOKUPS"]
    for i, s in enumerate(statuses, start=1):
        lookup_ws.cell(row=i, column=1, value=s)

    status_col = next(
        get_column_letter(i + 1)
        for i, (k, *_) in enumerate(FIELDS) if k == "workflow_status"
    )
    dv = DataValidation(
        type="list",
        formula1=f"_LOOKUPS!$A$1:$A${len(statuses)}",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid status",
        error="Choose a value from the dropdown list.",
    )
    ws.add_data_validation(dv)
    dv.sqref = f"{status_col}3:{status_col}500"

    return ws


def build_instructions_sheet(wb):
    ws = wb.create_sheet("INSTRUCTIONS")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 60

    headers = ["Field", "Required?", "Notes"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws.row_dimensions[1].height = 22

    for row_idx, (key, label, required, _, notes) in enumerate(FIELDS, start=2):
        ws.cell(row=row_idx, column=1, value=label).font = Font(name="Arial", size=10, bold=True)
        req_cell = ws.cell(row=row_idx, column=2, value="Yes ✓" if required else "No")
        req_cell.font = Font(
            name="Arial", size=10,
            color="C00000" if required else "595959"
        )
        ws.cell(row=row_idx, column=3, value=notes).font = Font(name="Arial", size=10)
        for col in range(1, 4):
            ws.cell(row=row_idx, column=col).border = border

    ws.append([])
    note_row = ws.max_row + 1
    ws.cell(row=note_row, column=1, value="NOTES").font = Font(bold=True, name="Arial")
    notes_text = [
        "• Yellow row 3 is an example — replace or delete it before importing.",
        "• Do NOT edit the _LOOKUPS sheet.",
        "• Dates must be YYYY-MM-DD format.",
        "• Run 1_discover.py first to populate releases and statuses.",
        "• Run 3_import_features.py only after filling this sheet.",
        "• The importer skips blank rows automatically.",
    ]
    for i, note in enumerate(notes_text, start=note_row + 1):
        ws.cell(row=i, column=1, value=note).font = Font(name="Arial", size=10)
        ws.merge_cells(f"A{i}:C{i}")


def main():
    summary = load_summary()

    wb = Workbook()

    # Create _LOOKUPS first (hidden)
    lookups_ws = wb.active
    lookups_ws.title = "_LOOKUPS"
    lookups_ws.sheet_state = "hidden"

    build_features_sheet(wb, summary)
    build_instructions_sheet(wb)

    # Reorder sheets: FEATURES first
    wb.move_sheet("FEATURES", offset=-wb.sheetnames.index("FEATURES"))

    output_path = "features_import.xlsx"
    wb.save(output_path)
    print(f"✅  Template created: {output_path}")
    print(f"    • {len(FIELDS)} columns ({sum(1 for _,_,r,_,_ in FIELDS if r)} required)")
    print(f"    • 47 data rows ready to fill (row 3 = example, rows 4–50 = yours)")
    print(f"    • Open in Excel or Numbers, fill rows, then run 3_import_features.py")


if __name__ == "__main__":
    main()
