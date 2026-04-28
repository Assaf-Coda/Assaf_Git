#!/usr/bin/env python3
"""
3_import_features.py
────────────────────
Reads features_import.xlsx and creates each feature in Aha! via the REST API.
Sends custom fields using the exact API keys discovered from the CODAV layout.
"""

import sys
import json
import time
import requests
import openpyxl
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from config import (
    AHA_API_KEY, AHA_SUBDOMAIN, PRODUCT_KEY,
    TARGET_RELEASE_NAME, STOP_ON_ERROR, SUPPRESS_NOTIFICATIONS
)

BASE_URL = f"https://{AHA_SUBDOMAIN}.aha.io/api/v1"
HEADERS = {
    "Authorization": f"Bearer {AHA_API_KEY}",
    "Content-Type":  "application/json",
    "Accept":        "application/json",
    "User-Agent":    f"aha-importer/1.0 ({AHA_SUBDOMAIN})",
}

INPUT_FILE   = "features_import.xlsx"
RESULTS_FILE = "import_results.xlsx"
DATA_START_ROW = 4
RATE_LIMIT_PAUSE = 0.4

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Column keys in exact order matching the template
COLUMN_KEYS = [
    "name", "type_allocation", "description", "conditions_of_satisfaction",
    "requires_enduser_documentation", "requires_ux", "assigned_to", "origin",
    "show_on_release_roadmap", "workflow_status", "tags", "start_date", "due_date",
]

# These are sent as top-level feature attributes
STANDARD_KEYS = {"name", "description", "workflow_status", "tags", "assigned_to",
                 "start_date", "due_date"}

# These are sent inside feature.custom_fields
CUSTOM_KEYS = {"type_allocation", "conditions_of_satisfaction",
               "requires_enduser_documentation", "requires_ux", "origin",
               "show_on_release_roadmap"}

# Required fields — importer validates before posting
REQUIRED_KEYS = {"name", "type_allocation", "description",
                 "requires_enduser_documentation", "requires_ux",
                 "assigned_to", "origin"}


def aha_get(path, params=None):
    resp = requests.get(f"{BASE_URL}{path}", headers=HEADERS, params=params)
    if resp.status_code == 401:
        print("❌  Authentication failed. Check AHA_API_KEY in config.py")
        sys.exit(1)
    resp.raise_for_status()
    return resp.json()


def resolve_release_ref():
    data = aha_get(f"/products/{PRODUCT_KEY}/releases", params={"per_page": 100})
    releases = data.get("releases", [])
    for r in releases:
        if r["name"].strip().lower() == TARGET_RELEASE_NAME.strip().lower():
            return r["reference_num"]
    names = [r["name"] for r in releases]
    print(f"❌  Release '{TARGET_RELEASE_NAME}' not found in {PRODUCT_KEY}.")
    print(f"    Available: {names}")
    sys.exit(1)


def validate_row(row_data, row_num):
    """Check required fields are filled. Returns error message or None."""
    missing = []
    for key in REQUIRED_KEYS:
        val = row_data.get(key, "").strip()
        if not val:
            missing.append(key)
    if missing:
        return f"Missing required fields: {', '.join(missing)}"
    return None


def create_feature(release_ref, row_data):
    payload = {"feature": {}}

    payload["feature"]["name"] = row_data["name"].strip()

    if row_data.get("description"):
        desc = row_data["description"].strip()
        if not desc.startswith("<"):
            desc = f"<p>{desc}</p>"
        payload["feature"]["description"] = desc

    if row_data.get("workflow_status"):
        payload["feature"]["workflow_status"] = {"name": row_data["workflow_status"]}

    if row_data.get("tags"):
        payload["feature"]["tags"] = row_data["tags"]

    if row_data.get("assigned_to"):
        payload["feature"]["assigned_to_user"] = {"email": row_data["assigned_to"].strip()}

    if row_data.get("start_date"):
        payload["feature"]["start_date"] = row_data["start_date"]

    if row_data.get("due_date"):
        payload["feature"]["due_date"] = row_data["due_date"]

    # Custom fields
    custom_fields = {}
    for key in CUSTOM_KEYS:
        val = row_data.get(key, "").strip()
        if val:
            custom_fields[key] = val

    if custom_fields:
        payload["feature"]["custom_fields"] = custom_fields

    if SUPPRESS_NOTIFICATIONS:
        payload["feature"]["send_email"] = False

    url = f"{BASE_URL}/releases/{release_ref}/features"
    try:
        resp = requests.post(url, headers=HEADERS, json=payload)
        if resp.status_code in (200, 201):
            feature = resp.json().get("feature", {})
            ref = feature.get("reference_num", "?")
            aha_url = feature.get("url", "")
            return True, f"Created {ref}", aha_url
        else:
            try:
                err = resp.json()
            except Exception:
                err = resp.text
            return False, f"HTTP {resp.status_code}: {err}", ""
    except requests.RequestException as e:
        return False, f"Request error: {e}", ""


def read_template():
    try:
        wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    except FileNotFoundError:
        print(f"❌  {INPUT_FILE} not found. Run 2_generate_template.py first.")
        sys.exit(1)

    ws = wb["FEATURES"]
    rows = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if all((v is None or str(v).strip() == "") for v in row):
            continue
        row_dict = {}
        for col_idx, key in enumerate(COLUMN_KEYS):
            val = row[col_idx] if col_idx < len(row) else None
            row_dict[key] = str(val).strip() if val is not None else ""
        rows.append(row_dict)
    return rows


def write_results(rows_with_status):
    wb = Workbook()
    ws = wb.active
    ws.title = "RESULTS"
    ws.freeze_panes = "A2"

    headers = ["#", "Feature Name", "Status", "Detail", "Aha! URL", "Timestamp"]
    widths  = [5, 45, 12, 50, 50, 22]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[chr(64 + col)].width = w

    ok_fill   = PatternFill("solid", start_color="E2EFDA")
    fail_fill = PatternFill("solid", start_color="FFDCE1")
    skip_fill = PatternFill("solid", start_color="F2F2F2")

    totals = {"created": 0, "failed": 0, "skipped": 0}

    for row_idx, item in enumerate(rows_with_status, start=2):
        success = item["success"]
        msg = item["message"]
        if success:
            label, fill = "✅ Created", ok_fill
            totals["created"] += 1
        elif "Skipped" in msg or "Missing" in msg:
            label, fill = "⏭ Skipped", skip_fill
            totals["skipped"] += 1
        else:
            label, fill = "❌ Failed", fail_fill
            totals["failed"] += 1

        for col, val in enumerate(
            [row_idx - 1, item["name"], label, msg, item["url"], item["timestamp"]],
            start=1
        ):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    sr = ws.max_row + 2
    ws.cell(row=sr, column=1,
            value=f"TOTAL: {totals['created']} created | {totals['failed']} failed | {totals['skipped']} skipped"
    ).font = Font(name="Arial", size=11, bold=True)
    ws.merge_cells(f"A{sr}:F{sr}")

    wb.save(RESULTS_FILE)
    return totals


def main():
    print("=" * 60)
    print("Aha! Feature Importer")
    print("=" * 60)

    print(f"🔍  Resolving release '{TARGET_RELEASE_NAME}' in {PRODUCT_KEY}...")
    release_ref = resolve_release_ref()
    print(f"✅  Found release: {release_ref}")

    print(f"📖  Reading {INPUT_FILE}...")
    rows = read_template()
    print(f"    {len(rows)} data rows found.\n")

    if not rows:
        print("⚠️  No rows to import.")
        sys.exit(0)

    # Pre-validate all rows
    print("🔎  Validating rows...")
    errors = []
    for i, row in enumerate(rows, start=1):
        err = validate_row(row, i)
        if err:
            errors.append((i, row.get("name", "(blank)"), err))

    if errors:
        print(f"\n⚠️  {len(errors)} row(s) have validation errors:\n")
        for num, name, err in errors:
            print(f"    Row {num}: {name[:40]} → {err}")
        print(f"\n    Fix these in {INPUT_FILE} and re-run.")
        sys.exit(1)

    print(f"    ✅  All {len(rows)} rows passed validation.\n")

    # ── Duplicate check: load existing feature names ─────────────────────
    print(f"🔎  Loading existing features from {PRODUCT_KEY} for duplicate check...")
    existing_names = set()
    page = 1
    while True:
        data = aha_get(f"/products/{PRODUCT_KEY}/features",
                       params={"per_page": 200, "page": page})
        feats = data.get("features", [])
        for f in feats:
            existing_names.add(f.get("name", "").strip().lower())
        total_pages = data.get("pagination", {}).get("total_pages", 1)
        if page >= total_pages:
            break
        page += 1
    print(f"    Found {len(existing_names)} existing features.\n")

    confirm = input(f"▶  Import {len(rows)} features into {PRODUCT_KEY} / {TARGET_RELEASE_NAME}? [y/N] ")
    if confirm.strip().lower() != "y":
        print("Cancelled.")
        sys.exit(0)

    print()
    results = []
    for i, row in enumerate(rows, start=1):
        name = row.get("name", "(blank)")
        print(f"  [{i:3d}/{len(rows)}] {name[:55]:<55}", end="  ")

        # Skip duplicates
        if name.strip().lower() in existing_names:
            success, message, aha_url = False, "Skipped — duplicate (already exists in Aha!)", ""
            print(f"⏭   {message}")
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            results.append({
                "success": success, "message": message,
                "url": aha_url, "name": name, "timestamp": ts
            })
            continue

        success, message, aha_url = create_feature(release_ref, row)
        # Add to existing set so within-batch duplicates are also caught
        if success:
            existing_names.add(name.strip().lower())
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        results.append({
            "success": success, "message": message,
            "url": aha_url, "name": name, "timestamp": ts
        })

        if success:
            print(f"✅  {message}")
        else:
            print(f"❌  {message}")
            if STOP_ON_ERROR:
                print("\n⛔  STOP_ON_ERROR — halting. Fix the issue and re-run.")
                break

        time.sleep(RATE_LIMIT_PAUSE)

    print()
    totals = write_results(results)
    print("=" * 60)
    print(f"  ✅  Created : {totals['created']}")
    print(f"  ❌  Failed  : {totals['failed']}")
    print(f"  ⏭   Skipped : {totals['skipped']}")
    print(f"\n  Results → {RESULTS_FILE}")
    print("=" * 60)


if __name__ == "__main__":
    main()