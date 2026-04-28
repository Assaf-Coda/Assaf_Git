#!/usr/bin/env python3
"""
2_generate_template.py
──────────────────────
Reads discovery_summary.json and generates features_import.xlsx.

Columns are built DYNAMICALLY from what the API discovered:
  • Standard fields: Name, Description, Workflow Status, Tags, etc.
  • Custom fields: each gets a column, with dropdowns for choice lists.

Sheet 1 — FEATURES      : one row per feature to import
Sheet 2 — INSTRUCTIONS  : field-by-field guidance
Sheet 3 — _LOOKUPS      : hidden data for dropdowns (do not edit)
"""

import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Colours ──────────────────────────────────────────────────────────────────
DARK_BLUE   = "1F4E79"
MID_BLUE    = "2E75B6"
GREY        = "595959"
WHITE       = "FFFFFF"
ROW_EVEN    = "EBF3FB"
ROW_ODD     = "FFFFFF"
EXAMPLE_BG  = "FFF2CC"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def load_summary():
    path = Path("discovery_summary.json")
    if not path.exists():
        print("❌  discovery_summary.json not found. Run 1_discover.py first.")
        sys.exit(1)
    with open(path) as f:
        return json.load(f)


# ── Column spec builder ─────────────────────────────────────────────────────

def build_columns(summary):
    """
    Returns a list of column dicts:
      { key, label, required, width, notes, is_custom, options }
    """
    cols = []

    # ── Standard fields ──────────────────────────────────────────────────
    cols.append({
        "key": "name", "label": "Feature Name *", "required": True,
        "width": 45, "notes": "Short, descriptive title",
        "is_custom": False, "options": [],
    })
    cols.append({
        "key": "description", "label": "Description *", "required": True,
        "width": 60,
        "notes": "User story format: As a <Actor>, I would like <Request>, so that <Value>.",
        "is_custom": False, "options": [],
    })
    cols.append({
        "key": "workflow_status", "label": "Workflow Status",
        "required": False, "width": 22,
        "notes": "Leave blank for default status",
        "is_custom": False,
        "options": summary.get("workflow_statuses", []),
    })
    cols.append({
        "key": "tags", "label": "Tags", "required": False, "width": 30,
        "notes": "Comma-separated, e.g. mobile,api",
        "is_custom": False, "options": [],
    })
    cols.append({
        "key": "assigned_to", "label": "Assigned To (email) *",
        "required": True, "width": 32,
        "notes": "Must be a valid Aha! user email address",
        "is_custom": False, "options": [],
    })
    cols.append({
        "key": "start_date", "label": "Start Date", "required": False,
        "width": 15, "notes": "YYYY-MM-DD",
        "is_custom": False, "options": [],
    })
    cols.append({
        "key": "due_date", "label": "Due Date", "required": False,
        "width": 15, "notes": "YYYY-MM-DD",
        "is_custom": False, "options": [],
    })

    # ── Custom fields from discovery ─────────────────────────────────────
    for cf in summary.get("custom_fields", []):
        # Skip if we already have a standard column for this
        if cf["key"] in ("name", "description", "workflow_status", "tags"):
            continue

        has_options = len(cf.get("options", [])) > 0
        label = cf["name"]
        cols.append({
            "key": cf["key"],
            "label": label,
            "required": False,  # we don't know from API; user will mark
            "width": max(20, min(40, len(label) + 6)),
            "notes": f"Custom field ({cf['type']})" + (
                f" — choose from dropdown" if has_options else ""
            ),
            "is_custom": True,
            "options": cf.get("options", []),
        })

    return cols


# ── Styling helpers ──────────────────────────────────────────────────────────

def style_header(cell, required):
    bg = DARK_BLUE if required else MID_BLUE
    cell.font = Font(bold=True, color=WHITE, size=11, name="Arial")
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border


def style_data(cell, row_idx):
    bg = ROW_EVEN if row_idx % 2 == 0 else ROW_ODD
    cell.fill = PatternFill("solid", start_color=bg)
    cell.font = Font(name="Arial", size=10)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = border


def style_example(cell):
    cell.fill = PatternFill("solid", start_color=EXAMPLE_BG)
    cell.font = Font(name="Arial", size=10, italic=True, color=GREY)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = border


# ── Sheet builders ───────────────────────────────────────────────────────────

def build_features_sheet(wb, columns, summary):
    ws = wb.active
    ws.title = "FEATURES"
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 36

    # Row 1 — banner
    ws.merge_cells(f"A1:{get_column_letter(len(columns))}1")
    banner = ws["A1"]
    banner.value = (
        f"Aha! Feature Import  ·  Product: {summary['product_key']}  ·  "
        f"Target Release: configured in config.py"
    )
    banner.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    banner.fill = PatternFill("solid", start_color=DARK_BLUE)
    banner.alignment = Alignment(horizontal="left", vertical="center")

    # Row 2 — headers
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col["label"])
        style_header(cell, col["required"])
        ws.column_dimensions[get_column_letter(col_idx)].width = col["width"]

    # Row 3 — example (yellow)
    example_values = {
        "name": "Support multi-tenant SSO login",
        "description": "As an enterprise admin, I would like SSO configuration, so that users can log in via their IdP.",
        "workflow_status": "",
        "tags": "security,enterprise",
        "assigned_to": "",
        "start_date": "2027-01-15",
        "due_date": "2027-03-31",
    }
    for col_idx, col in enumerate(columns, start=1):
        val = example_values.get(col["key"], "")
        # For custom fields with options, use first option as example
        if not val and col["options"]:
            val = col["options"][0] if col["options"] else ""
        cell = ws.cell(row=3, column=col_idx, value=val)
        style_example(cell)

    # Rows 4–55: blank data rows
    for row_idx in range(4, 56):
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx, value="")
            style_data(cell, row_idx)

    # ── _LOOKUPS sheet + dropdowns ───────────────────────────────────────
    lookup_ws = wb.create_sheet("_LOOKUPS")
    lookup_ws.sheet_state = "hidden"
    lookup_col = 1  # each options list gets its own column in _LOOKUPS

    for col_idx, col in enumerate(columns, start=1):
        if not col["options"]:
            continue

        # Write options into _LOOKUPS
        for opt_row, opt in enumerate(col["options"], start=1):
            lookup_ws.cell(row=opt_row, column=lookup_col, value=opt)

        # Create dropdown validation referencing _LOOKUPS
        lc = get_column_letter(lookup_col)
        fc = get_column_letter(col_idx)
        dv = DataValidation(
            type="list",
            formula1=f"_LOOKUPS!${lc}$1:${lc}${len(col['options'])}",
            allow_blank=True,
            showErrorMessage=True,
            errorTitle=f"Invalid {col['label']}",
            error="Choose a value from the dropdown list.",
        )
        ws.add_data_validation(dv)
        dv.sqref = f"{fc}3:{fc}500"

        lookup_col += 1

    return ws


def build_instructions_sheet(wb, columns):
    ws = wb.create_sheet("INSTRUCTIONS")
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 30

    headers = ["Field", "Required?", "Notes", "Type"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color=WHITE, name="Arial", size=11)
        cell.fill = PatternFill("solid", start_color=DARK_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 22

    for row_idx, col in enumerate(columns, start=2):
        ws.cell(row=row_idx, column=1, value=col["label"]).font = Font(
            name="Arial", size=10, bold=True
        )
        req_cell = ws.cell(
            row=row_idx, column=2,
            value="Yes ✓" if col["required"] else "No"
        )
        req_cell.font = Font(
            name="Arial", size=10,
            color="C00000" if col["required"] else GREY
        )
        ws.cell(row=row_idx, column=3, value=col["notes"]).font = Font(name="Arial", size=10)
        source = "Custom field" if col["is_custom"] else "Standard field"
        if col["options"]:
            source += f" (dropdown: {len(col['options'])} choices)"
        ws.cell(row=row_idx, column=4, value=source).font = Font(name="Arial", size=10)
        for c in range(1, 5):
            ws.cell(row=row_idx, column=c).border = border

    # Notes section
    ws.append([])
    note_row = ws.max_row + 1
    ws.cell(row=note_row, column=1, value="NOTES").font = Font(bold=True, name="Arial")
    notes = [
        "• Yellow row 3 is an example — replace or delete it before importing.",
        "• Do NOT edit the _LOOKUPS sheet.",
        "• Dates must be YYYY-MM-DD format.",
        "• Dark blue headers = required. Mid-blue = optional.",
        "• Custom fields with dropdowns are populated from your Aha! account.",
        "• Run 3_import_features.py after filling this sheet.",
    ]
    for i, note in enumerate(notes, start=note_row + 1):
        ws.cell(row=i, column=1, value=note).font = Font(name="Arial", size=10)
        ws.merge_cells(f"A{i}:D{i}")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    summary = load_summary()
    columns = build_columns(summary)

    wb = Workbook()
    build_features_sheet(wb, columns, summary)
    build_instructions_sheet(wb, columns)

    # Reorder: FEATURES first
    wb.move_sheet("FEATURES", offset=-wb.sheetnames.index("FEATURES"))

    output_path = "features_import.xlsx"
    wb.save(output_path)

    required_count = sum(1 for c in columns if c["required"])
    custom_count = sum(1 for c in columns if c["is_custom"])
    dropdown_count = sum(1 for c in columns if c["options"])

    print(f"✅  Template created: {output_path}")
    print(f"    • {len(columns)} columns total ({required_count} required)")
    print(f"    • {custom_count} custom fields from Aha! discovery")
    print(f"    • {dropdown_count} columns with dropdown validation")
    print(f"    • 52 data rows ready (row 3 = example, rows 4–55 = yours)")


if __name__ == "__main__":
    main()